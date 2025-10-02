import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
import os

class ExcelWriter:
    def write_to_excel(self, data, output_file):
        # Procesar los datos para separar texto y rutas de URL
        processed_data = []
        for row in data:
            new_row = row.copy()
            if 'URL' in new_row and isinstance(new_row['URL'], dict):
                # Extraer solo el texto para el DataFrame inicial
                new_row['URL'] = new_row['URL']['text']
            processed_data.append(new_row)
        
        df = pd.DataFrame(processed_data)
        
        # Reorganizar columnas para que URL aparezca al final
        columns = list(df.columns)
        if 'URL' in columns:
            columns.remove('URL')
            columns.append('URL')
            df = df[columns]
        
        # Guardar el DataFrame en Excel
        df.to_excel(output_file, index=False)
        
        # Cargar el archivo para ajustar el ancho de las columnas
        workbook = load_workbook(output_file)
        worksheet = workbook.active
        
        # Ajustar ancho de columnas específicas
        columns_to_adjust = ['B', 'C', 'D']  # Estilo, Description, Color
        column_names = ['Estilo', 'Description', 'Color']
        
        for col_letter, col_name in zip(columns_to_adjust, column_names):
            # Encontrar el ancho máximo del contenido
            max_length = 0
            for cell in worksheet[col_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # Establecer el ancho (añadir un poco de padding)
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            worksheet.column_dimensions[col_letter].width = adjusted_width
        
        # Ajustar el ancho de la columna URL (última columna)
        last_col_letter = chr(ord('A') + len(df.columns) - 1)
        max_length = 0
        for cell in worksheet[last_col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        # Establecer el ancho para la columna URL
        adjusted_width = min(max_length + 2, 80)  # Máximo 80 caracteres para URLs
        worksheet.column_dimensions[last_col_letter].width = adjusted_width
        
        # Crear hipervínculos en la columna URL
        url_column_index = len(df.columns)  # Índice de la columna URL (1-based)
        for row_idx, row_data in enumerate(data, start=2):  # Empezar desde la fila 2 (después del header)
            if 'URL' in row_data and isinstance(row_data['URL'], dict):
                cell = worksheet.cell(row=row_idx, column=url_column_index)
                file_path = row_data['URL']['path']
                link_text = row_data['URL']['text']
                
                # Crear hipervínculo usando el método más compatible
                if os.path.exists(file_path):
                    # Usar ruta absoluta directamente (funciona mejor en Windows)
                    cell.value = link_text
                    cell.hyperlink = file_path
                    cell.font = Font(color="0000FF", underline="single")  # Azul y subrayado
                else:
                    cell.value = f"{link_text} (archivo no encontrado)"
        
        # Guardar los cambios
        workbook.save(output_file)